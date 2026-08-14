# SPDX-FileCopyrightText: 2026 Tiago Alvarez Calderon Newton and SignalDojo Contributors
# SPDX-License-Identifier: GPL-3.0-or-later

from __future__ import annotations

from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

from app.campaign.comparison import compare_runs, detect_metric_outliers, export_comparison_tables
from app.campaign.execution import CampaignRunner
from app.exporters.campaign_report import (
    export_campaign_csv, export_campaign_excel, export_campaign_pdf, render_workflow_diagram_png,
    safe_spreadsheet_text, sanitise_sheet_name,
)
from tests.campaign_helpers import campaign, create_motor_files


def _executed_campaign(tmp_path: Path):
    create_motor_files(tmp_path)
    item = campaign(tmp_path)
    item.execution.detailed_result_limit = 20
    CampaignRunner(item).execute()
    item.reference_run_id = next(run.run_id for run in item.runs if "normal" in run.file_name)
    return item


def test_reference_comparison_metrics_signals_and_export(tmp_path: Path) -> None:
    item = _executed_campaign(tmp_path)
    selected = [run.run_id for run in item.runs if not run.errors][:3]
    result = compare_runs(item, selected, reference_run_id=item.reference_run_id, alignment="exact")
    assert len(result.metrics) == 3
    assert result.reference_run_id == item.reference_run_id
    assert result.signals
    reference_signal = next(signal for signal in result.signals if signal.run_id == item.reference_run_id)
    assert np.nanmax(np.abs(reference_signal.difference.values)) == pytest.approx(0.0)
    path = export_comparison_tables(result, tmp_path / "comparison.xlsx")
    assert path.exists() and openpyxl.load_workbook(path).sheetnames[0] == "Metrics"


def test_comparison_requires_explicit_alignment_for_different_time_bases(tmp_path: Path) -> None:
    item = _executed_campaign(tmp_path)
    selected_runs = [run for run in item.runs if not run.errors][:2]
    key = next(iter(selected_runs[1].detail_results))
    # Change one retained signal time base while preserving a valid serialized payload.
    from app.project.result_codec import deserialise_result, serialise_result
    from app.core.models import SignalData
    payload = selected_runs[1].detail_results[key]
    value = deserialise_result(payload)
    if not isinstance(value, SignalData):
        key = next(k for k, v in selected_runs[1].detail_results.items() if v.get("type") == "signal")
        value = deserialise_result(selected_runs[1].detail_results[key])
    shifted = value.with_values(value.values[:-1], time=value.time[:-1] + 0.0002, sample_rate=None)
    selected_runs[1].detail_results[key] = serialise_result(shifted)
    ids = [run.run_id for run in selected_runs]
    with pytest.raises(ValueError, match="different time bases"):
        compare_runs(item, ids, reference_run_id=ids[0], signal_key=key, alignment="exact")
    interpolated = compare_runs(item, ids, reference_run_id=ids[0], signal_key=key, alignment="overlap")
    assert len(interpolated.signals) == 2


def test_outlier_detection() -> None:
    frame = pd.DataFrame({"run_id": ["a", "b", "c", "d", "e"], "metric": [1.0, 1.1, 0.9, 1.05, 100.0]})
    assert detect_metric_outliers(frame)["metric"] == ["e"]


def test_campaign_report_exports_and_formula_injection(tmp_path: Path) -> None:
    item = _executed_campaign(tmp_path)
    item.campaign_metadata["operator_note"] = "=HYPERLINK(\"bad\")"
    item.runs[0].user_metadata["serial"] = "=1+1"
    csv_path = export_campaign_csv(item, tmp_path / "campaign.csv")
    excel_path = export_campaign_excel(item, tmp_path / "campaign.xlsx")
    pdf_path = export_campaign_pdf(item, tmp_path / "campaign.pdf", generated_utc="2026-01-01T00:00:00+00:00")
    assert csv_path.stat().st_size > 0 and pdf_path.stat().st_size > 1000
    assert "'=1+1" in csv_path.read_text(encoding="utf-8")
    workbook = openpyxl.load_workbook(excel_path, data_only=False)
    assert {"Summary", "Runs", "Metrics", "Requirements", "Errors", "Provenance"}.issubset(workbook.sheetnames)
    runs = workbook["Runs"]
    values = [cell.value for row in runs.iter_rows() for cell in row]
    assert "'=1+1" in values
    assert safe_spreadsheet_text("@SUM(A1:A2)").startswith("'")
    assert safe_spreadsheet_text("  =SUM(A1:A2)").startswith("'")
    assert len(render_workflow_diagram_png(item.workflow_document)) > 1000
    provenance_headers = [cell.value for cell in workbook["Provenance"][1]]
    assert "mapped_input_checksums" in provenance_headers and "workflow_version" in provenance_headers
    used: set[str] = set()
    first = sanitise_sheet_name("Bad/Name*", used); second = sanitise_sheet_name("Bad/Name*", used)
    assert first != second and len(first) <= 31 and len(second) <= 31


def test_campaign_report_bundle_can_be_cancelled_without_partial_files(tmp_path: Path) -> None:
    from app.exporters.campaign_report import CampaignReportCancelled, export_campaign_report_bundle

    item = campaign(tmp_path / "inputs")
    create_motor_files(Path(item.input_folder))
    CampaignRunner(item, project_directory=tmp_path).execute()
    output = tmp_path / "cancelled_reports"
    with pytest.raises(CampaignReportCancelled):
        export_campaign_report_bundle(item, output, is_cancelled=lambda: True)
    assert not list(output.glob("*"))
